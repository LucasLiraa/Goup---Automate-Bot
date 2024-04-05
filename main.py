import re
import requests
from io import BytesIO
from botbuilder.core import BotFrameworkAdapter, BotFrameworkAdapterSettings, TurnContext
from botbuilder.schema import Activity, ActivityTypes, ChannelAccount
from openpyxl import load_workbook
import streamlit as st

# Configurações do Bot Framework
APP_ID = 'f150075f-15bd-49b6-8192-9be8b8dfb01f'
SETTINGS = BotFrameworkAdapterSettings(APP_ID)
ADAPTER = BotFrameworkAdapter(SETTINGS)

# Função para consultar a planilha
async def consultar_planilha(USUÁRIO_LOGADO, NÚMERO_DO_SERIAL):
    # URL da planilha de consulta
    url_consulta = "https://hlbbrasil.sharepoint.com/:x:/r/sites/FS-MTZ-DEP/Documentos/TI/INFRAESTRUTURA/ADMINISTRA%C3%87%C3%83O/Milvus/2023/Relat%C3%B3rio%20personalizadoInventario_30-08-2023.xlsx?d=w0289c8a7b3874466a17ac6e0586fad46&csf=1&web=1&e=6CUpET"
    
    # Solicita HTTP para obter os dados da planilha de consulta
    response = requests.get(url_consulta)
    
    # Verifica se a solicitação foi bem-sucedida
    if response.status_code == 200:
        # Carrega a planilha de consulta
        wb_consulta = load_workbook(BytesIO(response.content))
        ws_consulta = wb_consulta.active
        
        # Iterando sobre as linhas da planilha para encontrar a correspondência
        for row in ws_consulta.iter_rows(min_row=2):
            if USUÁRIO_LOGADO and row[0].value == USUÁRIO_LOGADO:
                return row[1].value, row[2].value, row[3].value  # Retornando os dados correspondentes
            elif NÚMERO_DO_SERIAL and row[1].value == NÚMERO_DO_SERIAL:
                return row[0].value, row[2].value, row[3].value  # Retornando os dados correspondentes

    # Se não encontrou correspondência, retorna None para cada campo
    return "", "", ""

# Função para adicionar à planilha
async def adicionar_a_planilha(USUÁRIO_LOGADO, NÚMERO_DO_SERIAL, OBSERVAÇÃO_DO_DISPOSITIVO, MEMÓRIA_RAM_TOTAL):
    # URL da planilha de estoque
    url_estoque = "https://hlbbrasil.sharepoint.com/:x:/r/sites/FS-MTZ-DEP/Documentos/TI/INFRAESTRUTURA/AREA%20COMUM/LUCAS%20LIRA/Estoque%20atual.xlsx?d=weef5bfdf7f8448d6b01d8bb90749754f&csf=1&web=1&e=wGPqhr"
    
    # Solicita HTTP para obter os dados da planilha de estoque
    response = requests.get(url_estoque)
    
    # Verifica se a solicitação foi bem-sucedida
    if response.status_code == 200:
        # Carrega a planilha de estoque
        wb_estoque = load_workbook(BytesIO(response.content))
        ws_estoque = wb_estoque.active
        
        # Encontra a próxima linha vazia na planilha de estoque
        next_row = ws_estoque.max_row + 1
        
        # Adiciona as informações à planilha de estoque
        ws_estoque.cell(row=next_row, column=1).value = USUÁRIO_LOGADO
        ws_estoque.cell(row=next_row, column=2).value = NÚMERO_DO_SERIAL
        ws_estoque.cell(row=next_row, column=3).value = OBSERVAÇÃO_DO_DISPOSITIVO
        ws_estoque.cell(row=next_row, column=4).value = MEMÓRIA_RAM_TOTAL
        
        # Salva as alterações
        wb_estoque.save(url_estoque)

# Função para lidar com a atividade de mensagem
async def on_message_activity(turn_context: TurnContext):
    if turn_context.activity.type == ActivityTypes.message:
        text = turn_context.activity.text
        pattern = r'"(.*?)"'
        matches = re.findall(pattern, text)

        if matches:
            USUÁRIO_LOGADO = matches[0] if len(matches) >= 1 else None
            NÚMERO_DO_SERIAL = matches[1] if len(matches) >= 2 else None

            # Consulta a planilha online
            USUÁRIO_LOGADO_found, OBSERVAÇÃO_DO_DISPOSITIVO, MEMÓRIA_RAM_TOTAL = await consultar_planilha(USUÁRIO_LOGADO, NÚMERO_DO_SERIAL)

            # Responde com as informações encontradas
            if USUÁRIO_LOGADO_found:
                reply_text = f"Equipamento adicionado ao estoque."
                
                # Adiciona as informações à planilha de estoque
                await adicionar_a_planilha(USUÁRIO_LOGADO_found, NÚMERO_DO_SERIAL, OBSERVAÇÃO_DO_DISPOSITIVO, MEMÓRIA_RAM_TOTAL)
            else:
                reply_text = "Desculpe, não encontrei o equipamento."
        else:
            reply_text = "Formato da mensagem inválido."

        # Envia resposta ao remetente
        reply_activity = Activity(
            type=ActivityTypes.message,
            text=reply_text,
            recipient=ChannelAccount(id=turn_context.activity.from_property.id),
            from_property=ChannelAccount(id=turn_context.activity.recipient.id),
        )
        await turn_context.send_activity(reply_activity)

# Função principal para processar a solicitação
async def main(req):
    turn_context = await ADAPTER.create_turn_context(req)
    await ADAPTER.process_activity(turn_context, on_message_activity)

# Configuração do Streamlit
st.title("Código do Bot")
st.write("Aqui está o código do bot:")
st.code("""
import re
import requests
from io import BytesIO
from botbuilder.core import BotFrameworkAdapter, BotFrameworkAdapterSettings, TurnContext
from botbuilder.schema import Activity, ActivityTypes, ChannelAccount
from openpyxl import load_workbook

# Código do bot aqui...

if __name__ == "__main__":
    # Configuração do bot...
""")
